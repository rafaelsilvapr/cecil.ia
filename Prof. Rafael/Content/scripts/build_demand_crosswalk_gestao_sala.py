from __future__ import annotations

import csv
import shutil
from collections import defaultdict
from pathlib import Path

from openpyxl import Workbook, load_workbook


BASE_DIR = Path("/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael")
DRIVE_BASE_DIR = Path(
    "/Users/rafaelrodriguesdasilva/Library/CloudStorage/GoogleDrive-rafaelsilva.pr@gmail.com/Meu Drive/Empreendedor/Professor Rafael/Prof. Rafael"
)

TRENDS_CSV = BASE_DIR / "output/spreadsheet/mapa_mercado_ebooks_gestao_sala_google_trends_2026-04-08.csv"
YOUTUBE_SOURCE_XLSX = BASE_DIR / "Bkp_Fase3_Antigos/Relatorio_Analise_YouTube_Rafael_backup_pre_comments.xlsx"

LOCAL_SHEET_DIR = BASE_DIR / "output/spreadsheet"
DRIVE_SHEET_DIR = DRIVE_BASE_DIR / "output/spreadsheet"

LOCAL_RAW_DIR = BASE_DIR / "Content/base_de_dados/ebooks_gestao_sala/raw/youtube_comment_signals/2026-04-08"
DRIVE_RAW_DIR = DRIVE_BASE_DIR / "Content/base_de_dados/ebooks_gestao_sala/raw/youtube_comment_signals/2026-04-08"

LOCAL_RAW_PARENT = LOCAL_RAW_DIR.parent
DRIVE_RAW_PARENT = DRIVE_RAW_DIR.parent

SYNTHESIS_CSV = "mapa_mercado_ebooks_gestao_sala_demand_crosswalk_2026-04-08.csv"
SYNTHESIS_XLSX = "mapa_mercado_ebooks_gestao_sala_demand_crosswalk_2026-04-08.xlsx"
RAW_SIGNALS_CSV = "youtube_comment_signals_2026-04-08.csv"
MANIFEST_NAME = "2026-04-08_manifest.md"


THEMES = [
    {
        "theme_id": "tempo_planejamento",
        "theme": "Tempo, planejamento e quick wins",
        "trend_terms": ["gestão de sala de aula"],
        "signal_texts": [
            "Efficiency, automation, and time-saving using simple rules.",
        ],
        "declared_pain": "Nao consigo planejar tudo e preciso de uma aula pronta sem perder qualidade.",
        "latent_pain": "Sobrecarga mental, culpa e sensacao de estar sempre correndo atras.",
        "purchase_language": "30 minutos, quick-win, guia rapido, automacao, regras simples.",
        "book_implication": "Abrir o livro com alavancas de economia de tempo e sequencias prontas.",
        "confidence": "Alta",
        "book_use": "Capitulo central e promessa de capa",
    },
    {
        "theme_id": "indisciplina_controle",
        "theme": "Indisciplina e recuperacao de controle",
        "trend_terms": ["indisciplina", "controle de sala de aula"],
        "signal_texts": [
            "Identification with daily pain points / Early childhood struggles.",
        ],
        "declared_pain": "A turma desafia meus limites e eu sinto que perdi a sala.",
        "latent_pain": "Quero recuperar autoridade sem virar autoritario e sem me esgotar.",
        "purchase_language": "retomar controle, comportamento, estrategia, gestao de sala, autoridade pratica.",
        "book_implication": "Capitulo de abertura forte para dor central do mercado.",
        "confidence": "Alta",
        "book_use": "Promessa principal e entrada do conteudo",
    },
    {
        "theme_id": "protocolos_limites",
        "theme": "Regras, limites e protocolos",
        "trend_terms": ["disciplina escolar", "controle de sala de aula"],
        "signal_texts": [
            "Requests for 'protocol' PDF guide / parental privacy rights.",
            "School protocols for phone-free environments.",
        ],
        "declared_pain": "Preciso de um protocolo claro para agir em situacoes delicadas.",
        "latent_pain": "Quero respaldo e previsibilidade para nao improvisar sob pressao.",
        "purchase_language": "protocolo PDF, checklist, guia, regras prontas, contrato, celular.",
        "book_implication": "Bonus, anexos ou capitulo tatico com modelos prontos.",
        "confidence": "Alta",
        "book_use": "Anexos, checklists e materiais de apoio",
    },
    {
        "theme_id": "scripts_conversas",
        "theme": "Scripts para conversas dificeis",
        "trend_terms": ["indisciplina", "controle de sala de aula"],
        "signal_texts": [
            "Scripts for difficult workplace conversations for educators.",
        ],
        "declared_pain": "Nao sei o que dizer na hora do conflito e travo diante da conversa.",
        "latent_pain": "Tenho medo de piorar a situacao e perder a autoridade.",
        "purchase_language": "scripts prontos, frases prontas, o que dizer, roteiro.",
        "book_implication": "Capitulo muito forte para familia-escola e conflitos em sala.",
        "confidence": "Alta",
        "book_use": "Capitulo pratico e material de apoio",
    },
    {
        "theme_id": "inicio_carreira_sobrecarga",
        "theme": "Inicio de carreira e sobrecarga",
        "trend_terms": ["gestão de sala de aula", "indisciplina"],
        "signal_texts": [
            "HUGE: Quick-win grammar guides for busy teachers.",
        ],
        "declared_pain": "Sou iniciante e preciso sobreviver aos primeiros dias sem me perder.",
        "latent_pain": "Tenho medo de comecar errado, ser engolido pelo caos e me sentir incapaz.",
        "purchase_language": "manual, guia de sobrevivencia, primeiros 90 dias, para professores iniciantes.",
        "book_implication": "Esse recorte e forte para titulo, subtitulo e definicao de publico.",
        "confidence": "Alta",
        "book_use": "Titulo secundario, subtitulo e posicionamento do produto",
    },
    {
        "theme_id": "neurodivergencia_tdaH",
        "theme": "Neurodivergencia e TDAH",
        "trend_terms": ["manejo de classe", "indisciplina"],
        "signal_texts": [
            "Specific requests for TDAH in Adults / Roadmaps.",
        ],
        "declared_pain": "Quero lidar com TDAH e diferencas sem perder a conducao da turma.",
        "latent_pain": "Preciso adaptar sem quebrar a rotina nem me sentir impotente.",
        "purchase_language": "roadmap, TDAH, guia, como agir.",
        "book_implication": "Melhor como capitulo complementar ou bonus especializado.",
        "confidence": "Media",
        "book_use": "Bonus ou capitulo adjacente",
    },
]


def read_trends_summary(path: Path) -> dict[str, dict]:
    with path.open(newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))

    grouped: dict[str, list[dict]] = defaultdict(list)
    for row in rows:
        grouped[row["term"]].append(row)

    aggregated: dict[str, dict] = {}
    for term, items in grouped.items():
        positive = [r for r in items if int(r["max_interest"]) > 0]
        aggregated[term] = {
            "term": term,
            "source_files": len({r["source_file_id"] for r in positive}),
            "max_interest": max((int(r["max_interest"]) for r in positive), default=0),
            "first_nonzero_date": min(
                (r["first_nonzero_date"] for r in positive if r["first_nonzero_date"] != "nao publico"),
                default="nao publico",
            ),
            "nonzero_months_min": min((int(r["nonzero_months"]) for r in positive), default=0),
            "nonzero_months_max": max((int(r["nonzero_months"]) for r in positive), default=0),
            "search_intent": items[0]["search_intent"],
            "pain_inferida": items[0]["pain_inferida"],
            "confidence": items[0]["confidence"],
        }
    return aggregated


def extract_youtube_signals(path: Path) -> list[dict]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb["2) Conteudo Viral"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    rows: list[dict] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        signal = row[idx["Demand Signals"]]
        if not signal:
            continue
        rows.append(
            {
                "channel": row[idx["Canal"]],
                "video_title": row[idx["Video Title"]],
                "views": row[idx["Views"]],
                "comments_per_1k": row[idx["Comments / 1k views"]],
                "likes_per_1k": row[idx["Likes / 1k views"]],
                "demand_signal": signal,
                "outlier_score": row[idx["Outlier Score"]],
                "data_coleta": row[idx["Data da coleta"]],
            }
        )
    return rows


def signal_to_theme_map() -> dict[str, str]:
    mapping = {}
    for theme in THEMES:
        for signal in theme["signal_texts"]:
            mapping[signal] = theme["theme_id"]
    return mapping


def build_synthesis_rows(trends: dict[str, dict], signals: list[dict]) -> list[dict]:
    theme_lookup = {theme["theme_id"]: theme for theme in THEMES}
    mapped_signals: dict[str, list[dict]] = defaultdict(list)
    sig_map = signal_to_theme_map()
    for signal in signals:
        mapped = sig_map.get(signal["demand_signal"])
        if mapped:
            mapped_signals[mapped].append(signal)

    rows: list[dict] = []
    for theme in THEMES:
        trend_terms = [term for term in theme["trend_terms"] if term in trends]
        trend_summary = []
        for term in trend_terms:
            t = trends[term]
            trend_summary.append(
                f"{term} (pico {t['max_interest']}, {t['source_files']} arquivo(s), primeira aparicao {t['first_nonzero_date']})"
            )
        youtube_summary = [
            f"{s['channel']} | {s['video_title']} -> {s['demand_signal']}" for s in mapped_signals.get(theme["theme_id"], [])
        ]
        rows.append(
            {
                "theme_id": theme["theme_id"],
                "tema": theme["theme"],
                "google_trends_terms": " ; ".join(trend_summary) if trend_summary else "nao mapeado",
                "youtube_signals": " ; ".join(youtube_summary) if youtube_summary else "nao mapeado",
                "declared_pain": theme["declared_pain"],
                "latent_pain": theme["latent_pain"],
                "purchase_language": theme["purchase_language"],
                "fit_to_book": theme["book_use"],
                "confidence": theme["confidence"],
                "book_implication": theme["book_implication"],
            }
        )
    return rows


def build_youtube_rows(signals: list[dict]) -> list[dict]:
    sig_map = signal_to_theme_map()
    theme_lookup = {theme["theme_id"]: theme for theme in THEMES}
    out: list[dict] = []
    for signal in signals:
        theme_id = sig_map.get(signal["demand_signal"], "nao_mapeado")
        theme = theme_lookup.get(theme_id)
        out.append(
            {
                "channel": signal["channel"],
                "video_title": signal["video_title"],
                "views": signal["views"],
                "comments_per_1k": signal["comments_per_1k"],
                "likes_per_1k": signal["likes_per_1k"],
                "demand_signal": signal["demand_signal"],
                "mapped_theme": theme["theme"] if theme else "nao mapeado",
                "declared_pain": theme["declared_pain"] if theme else "",
                "latent_pain": theme["latent_pain"] if theme else "",
                "purchase_language": theme["purchase_language"] if theme else "",
                "confidence": theme["confidence"] if theme else "",
                "source_note": "Extraido do campo Demand Signals em Relatorio_Analise_YouTube_Rafael_backup_pre_comments.xlsx",
            }
        )
    return out


def build_trends_rows(trends: dict[str, dict]) -> list[dict]:
    role_map = {
        "indisciplina": "declared_pain",
        "disciplina escolar": "declared_pain",
        "gestão de sala de aula": "latent_pain",
        "controle de sala de aula": "purchase_language",
        "manejo de classe": "academic_adjacent",
    }
    rows: list[dict] = []
    for term, info in trends.items():
        rows.append(
            {
                "term": term,
                "signal_role": role_map.get(term, "nao_mapeado"),
                "files_with_signal": info["source_files"],
                "max_interest": info["max_interest"],
                "first_nonzero_date": info["first_nonzero_date"],
                "nonzero_months_min": info["nonzero_months_min"],
                "nonzero_months_max": info["nonzero_months_max"],
                "search_intent": info["search_intent"],
                "pain_inferida": info["pain_inferida"],
                "confidence": info["confidence"],
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheets: dict[str, tuple[list[dict], list[str]]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    first = True
    for sheet_name, (rows, fieldnames) in sheets.items():
        ws = wb.active if first else wb.create_sheet(title=sheet_name)
        ws.title = sheet_name
        ws.freeze_panes = "A2"
        ws.append(fieldnames)
        for row in rows:
            ws.append([row.get(col, "") for col in fieldnames])
        first = False
    wb.save(path)


def write_text(path: Path, text: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text, encoding="utf-8")


def make_manifest(trends_rows: list[dict], youtube_rows: list[dict], synthesis_rows: list[dict]) -> str:
    lines = [
        "# Crosswalk de Demanda - Gestao de Sala de Aula",
        "",
        "Dataset date: 2026-04-08",
        "",
        "## Como ler",
        "- Google Trends entra como sinal de busca declarada e recorrente.",
        "- O eixo de YouTube entra como sinal extraido de comentarios e de linguagem de engajamento/pedido.",
        "- A sintese separa dor declarada, dor latente e linguagem de compra para orientar titulo, subtitulo e promessa.",
        "",
        "## O que ficou mais forte",
        "- Indisciplina continua como a dor central mais dura e mais duradoura.",
        "- Tempo, planejamento e quick wins surgem como a segunda alavanca mais clara.",
        "- Protocolos, scripts e frases prontas aparecem como linguagem de compra muito forte.",
        "- O recorte de professor iniciante entra como forte promessa de acolhimento e sobrevivencia.",
        "",
        "## Arquivos registrados",
        f"- Trends agregados: {len(trends_rows)} termos",
        f"- Sinais do YouTube: {len(youtube_rows)} registros",
        f"- Sintese de temas: {len(synthesis_rows)} temas",
    ]
    return "\n".join(lines)


def main() -> int:
    trends = read_trends_summary(TRENDS_CSV)
    youtube_signals = extract_youtube_signals(YOUTUBE_SOURCE_XLSX)
    synthesis_rows = build_synthesis_rows(trends, youtube_signals)
    trends_rows = build_trends_rows(trends)
    youtube_rows = build_youtube_rows(youtube_signals)

    synthesis_fields = [
        "theme_id",
        "tema",
        "google_trends_terms",
        "youtube_signals",
        "declared_pain",
        "latent_pain",
        "purchase_language",
        "fit_to_book",
        "confidence",
        "book_implication",
    ]
    trends_fields = [
        "term",
        "signal_role",
        "files_with_signal",
        "max_interest",
        "first_nonzero_date",
        "nonzero_months_min",
        "nonzero_months_max",
        "search_intent",
        "pain_inferida",
        "confidence",
    ]
    youtube_fields = [
        "channel",
        "video_title",
        "views",
        "comments_per_1k",
        "likes_per_1k",
        "demand_signal",
        "mapped_theme",
        "declared_pain",
        "latent_pain",
        "purchase_language",
        "confidence",
        "source_note",
    ]

    for sheet_dir in (LOCAL_SHEET_DIR, DRIVE_SHEET_DIR):
        sheet_dir.mkdir(parents=True, exist_ok=True)
        write_csv(sheet_dir / SYNTHESIS_CSV, synthesis_rows, synthesis_fields)
        write_xlsx(
            sheet_dir / SYNTHESIS_XLSX,
            {
                "Sintese": (synthesis_rows, synthesis_fields),
                "Trends_Termos": (trends_rows, trends_fields),
                "Youtube_Sinais": (youtube_rows, youtube_fields),
            },
        )

    manifest_text = make_manifest(trends_rows, youtube_rows, synthesis_rows)
    raw_readme = "\n".join(
        [
            "# YouTube Comment Signals - Gestao de Sala de Aula",
            "",
            "Pasta com sinais extraidos do campo Demand Signals da planilha de auditoria de YouTube.",
            "",
            "## Cuidado de leitura",
            "- Nao trata isso como transcricao literal de comentarios brutos.",
            "- Usa os sinais como derivacao analitica de pedidos, dores e linguagem de compra.",
            "",
            "## Origem",
            "- Fonte principal: Relatorio_Analise_YouTube_Rafael_backup_pre_comments.xlsx, aba 2) Conteudo Viral.",
            "- Os sinais sao cruzados com Google Trends para separar dor declarada, dor latente e linguagem de compra.",
        ]
    )

    for raw_dir in (LOCAL_RAW_DIR, DRIVE_RAW_DIR):
        raw_dir.mkdir(parents=True, exist_ok=True)
        write_csv(raw_dir / RAW_SIGNALS_CSV, youtube_rows, youtube_fields)
        write_text(raw_dir / "README.md", raw_readme)

    for parent in (LOCAL_RAW_PARENT, DRIVE_RAW_PARENT):
        parent.mkdir(parents=True, exist_ok=True)
        write_text(parent / MANIFEST_NAME, manifest_text)

    print(f"wrote {len(synthesis_rows)} synthesis rows and {len(youtube_rows)} youtube signal rows")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
