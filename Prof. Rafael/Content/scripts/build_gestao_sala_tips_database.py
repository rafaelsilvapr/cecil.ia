from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo


BASE_DIR = Path("/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael")
DRIVE_BASE_DIR = Path(
    "/Users/rafaelrodriguesdasilva/Library/CloudStorage/GoogleDrive-rafaelsilva.pr@gmail.com/Meu Drive/Empreendedor/Professor Rafael/Prof. Rafael"
)

TARGET_ROOTS = [BASE_DIR, DRIVE_BASE_DIR]

CSV_REL = Path("output/spreadsheet/mapa_dicas_gestao_sala_2026-04-09.csv")
XLSX_REL = Path("output/spreadsheet/mapa_dicas_gestao_sala_2026-04-09.xlsx")
README_REL = Path("Content/base_de_dados/ebooks_gestao_sala/base_dicas/README.md")

SLOT_COUNT = 60

CSV_HEADERS = [
    "slot_id",
    "tip_texto",
    "tema",
    "subtema",
    "dor_associada",
    "novidade_0a10",
    "impacto_0a10",
    "evidencia_0a10",
    "conforto_publico_0a10",
    "risco_cliche_0a10",
    "score_total",
    "prioridade",
    "uso_ideal",
    "status",
    "fonte_ids",
    "fonte_principal",
    "trecho_prova",
    "observacoes",
]

SOURCE_ROWS = [
    [
        "F01",
        "dado primario",
        "YouTube comment signals",
        "/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael/Content/base_de_dados/ebooks_gestao_sala/raw/youtube_comment_signals/2026-04-08/youtube_comment_signals_2026-04-08.csv",
        "validado",
        "Base de comentarios e dores extraidas de videos de educacao.",
        "Usar para validar linguagem de dor, rotina, sobrecarga e primeiro emprego.",
    ],
    [
        "F02",
        "dado primario",
        "Google Trends exports",
        "/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael/Content/base_de_dados/ebooks_gestao_sala/raw/google_trends/2026-04-08_manifest.md",
        "validado",
        "Sinal de demanda temporal e recorrencia de termos.",
        "Usar para separar linguagem de busca de linguagem de venda.",
    ],
    [
        "F03",
        "dado primario",
        "Sales pages and copy texts",
        "/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael/Content/base_de_dados/ebooks_gestao_sala/raw/copy_texts/",
        "validado",
        "Copys publicas de e-books, cursos e paginas de venda.",
        "Usar para estudar promessa, titulo, mecanismo e prova social.",
    ],
    [
        "F04",
        "dado primario",
        "Sales fronts manifest",
        "/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael/Content/base_de_dados/ebooks_gestao_sala/raw/sales_pages/2026-04-08_manifest.md",
        "validado",
        "Registro das frentes de venda catalogadas no projeto.",
        "Usar para cruzar ebook, curso e promessa comercial.",
    ],
    [
        "F05",
        "dado secundario",
        "Market summary",
        "/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael/Content/base_de_dados/ebooks_gestao_sala/resumo_mercado_2026-04-08.md",
        "validado",
        "Resumo consolidado das leituras de mercado.",
        "Usar como sintese de leitura do nicho.",
    ],
    [
        "F06",
        "externo",
        "Gemini shared report",
        "https://gemini.google.com/share/629a52940a0e",
        "pendente",
        "Relatorio compartilhado pelo usuario, ainda nao acessivel neste ambiente.",
        "Importar quando o texto bruto estiver disponivel.",
    ],
]

COLUMN_WIDTHS = {
    "A": 10,
    "B": 44,
    "C": 18,
    "D": 20,
    "E": 22,
    "F": 12,
    "G": 12,
    "H": 12,
    "I": 16,
    "J": 14,
    "K": 11,
    "L": 10,
    "M": 18,
    "N": 18,
    "O": 18,
    "P": 34,
    "Q": 34,
    "R": 28,
}


def rel_path(root: Path, rel: Path) -> Path:
    return root / rel


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def make_rows() -> list[dict[str, str]]:
    rows: list[dict[str, str]] = []
    for index in range(1, SLOT_COUNT + 1):
        rows.append(
            {
                "slot_id": f"D{index:02d}",
                "tip_texto": "",
                "tema": "",
                "subtema": "",
                "dor_associada": "",
                "novidade_0a10": "",
                "impacto_0a10": "",
                "evidencia_0a10": "",
                "conforto_publico_0a10": "",
                "risco_cliche_0a10": "",
                "score_total": "",
                "prioridade": "",
                "uso_ideal": "",
                "status": "aguardando_importacao",
                "fonte_ids": "",
                "fonte_principal": "",
                "trecho_prova": "",
                "observacoes": "",
            }
        )
    return rows


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=CSV_HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def style_header(ws, max_col: int) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    font = Font(color="FFFFFF", bold=True)
    border = Border(bottom=Side(style="thin", color="D9E2F3"))
    for col_idx in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = fill
        cell.font = font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def apply_body_style(ws, start_row: int, end_row: int, max_col: int) -> None:
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: dict[str, int]) -> None:
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def create_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    ensure_parent(path)

    wb = Workbook()
    ws = wb.active
    ws.title = "Dicas"

    ws.append(CSV_HEADERS)
    for row in rows:
        ws.append([row[h] for h in CSV_HEADERS])

    # Formula columns in the workbook version. They stay blank until the score cells are filled.
    for row_idx in range(2, SLOT_COUNT + 2):
        ws[f"K{row_idx}"] = f'=IF(COUNTA(F{row_idx}:J{row_idx})=0,"",ROUND((F{row_idx}+G{row_idx}+H{row_idx}+I{row_idx}+(10-J{row_idx}))/5,1))'
        ws[f"L{row_idx}"] = f'=IF(K{row_idx}="","",IF(K{row_idx}>=8,"A",IF(K{row_idx}>=6,"B","C")))'

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{SLOT_COUNT + 1}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 24

    style_header(ws, len(CSV_HEADERS))
    apply_body_style(ws, 2, SLOT_COUNT + 1, len(CSV_HEADERS))
    set_widths(ws, COLUMN_WIDTHS)

    # Data validation
    score_validation = DataValidation(type="whole", operator="between", formula1="0", formula2="10", allow_blank=True)
    score_validation.prompt = "Use um numero inteiro de 0 a 10."
    score_validation.error = "Valor invalido. Use um numero inteiro de 0 a 10."
    ws.add_data_validation(score_validation)
    score_validation.add(f"F2:J{SLOT_COUNT + 1}")

    status_validation = DataValidation(
        type="list",
        formula1='"aguardando_importacao,em_analise,validado,descartado"',
        allow_blank=True,
    )
    ws.add_data_validation(status_validation)
    status_validation.add(f"N2:N{SLOT_COUNT + 1}")

    uso_validation = DataValidation(
        type="list",
        formula1='"reels,youtube_curto,youtube_longo,ebook,carrossel,newsletter"',
        allow_blank=True,
    )
    ws.add_data_validation(uso_validation)
    uso_validation.add(f"M2:M{SLOT_COUNT + 1}")

    tab = Table(displayName="DicasTable", ref=f"A1:R{SLOT_COUNT + 1}")
    tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(tab)

    # Sources sheet
    src = wb.create_sheet("Fontes")
    src_headers = ["fonte_id", "categoria", "titulo", "url_ou_path", "status", "descricao", "observacoes"]
    src.append(src_headers)
    for row in SOURCE_ROWS:
        src.append(row)
    src.freeze_panes = "A2"
    src.auto_filter.ref = f"A1:G{len(SOURCE_ROWS) + 1}"
    src.sheet_view.showGridLines = False
    style_header(src, len(src_headers))
    apply_body_style(src, 2, len(SOURCE_ROWS) + 1, len(src_headers))
    src_widths = {"A": 10, "B": 16, "C": 24, "D": 52, "E": 12, "F": 42, "G": 34}
    set_widths(src, src_widths)
    src_tab = Table(displayName="FontesTable", ref=f"A1:G{len(SOURCE_ROWS) + 1}")
    src_tab.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    src.add_table(src_tab)

    # Criteria sheet
    crit = wb.create_sheet("Criterios")
    crit_headers = ["campo", "faixa_baixa", "faixa_media", "faixa_alta", "faixa_maxima", "regra_de_selecao"]
    crit.append(crit_headers)
    crit_rows = [
        [
            "novidade",
            "ja existe em quase tudo",
            "parecido com o mercado",
            "tem um angulo novo",
            "abre espaco real de diferenca",
            "A novidade conta mais quando nao destrói a clareza.",
        ],
        [
            "impacto",
            "resolve pouco",
            "ajuda, mas e parcial",
            "resolve uma dor forte",
            "muda o dia a dia rapidamente",
            "Impacto alto vale mais do que criatividade vazia.",
        ],
        [
            "evidencia",
            "isolada ou fraca",
            "aparece em uma fonte",
            "aparece em mais de uma fonte",
            "repete em fontes diferentes e confiaveis",
            "A melhor ideia junta evidencia de oferta, busca e fala do publico.",
        ],
        [
            "conforto_publico",
            "soa acusatorio",
            "soa duro, mas suportavel",
            "soa profissional e humano",
            "soa acolhedor e util",
            "Evite linguagem que faça o professor sentir fracasso sem saida.",
        ],
        [
            "risco_cliche",
            "muito repetido e vazio",
            "comum e pouco diferencial",
            "convencional, mas aceitavel",
            "fora do lugar-comum",
            "Cliche so entra quando a prova e muito forte e o mecanismo e concreto.",
        ],
        [
            "regra_final",
            "exploratorio",
            "em observacao",
            "prioridade media",
            "prioridade alta",
            "Prefira ideias com evidencia alta e risco de cliche controlado.",
        ],
    ]
    for row in crit_rows:
        crit.append(row)
    crit.freeze_panes = "A2"
    crit.auto_filter.ref = f"A1:F{len(crit_rows) + 1}"
    crit.sheet_view.showGridLines = False
    style_header(crit, len(crit_headers))
    apply_body_style(crit, 2, len(crit_rows) + 1, len(crit_headers))
    set_widths(crit, {"A": 18, "B": 24, "C": 24, "D": 24, "E": 24, "F": 44})

    wb.save(path)


def write_readme(path: Path) -> None:
    ensure_parent(path)
    content = """# Base de dicas - Gestao de sala de aula

Esta pasta e a planilha associada organizam a futura base de dicas, insights e ensinamentos para Reels, videos longos e proximos ebooks.

## Objetivo
- Guardar cada dica como um item independente.
- Classificar cada item por novidade, impacto, conforto para o professor e risco de clichê.
- Registrar as fontes que sustentam cada dica.

## Estrutura da planilha
- `Dicas`: 60 slots prontos para importar e classificar as dicas.
- `Fontes`: registro das fontes e corpus usados para sustentar cada item.
- `Criterios`: guia de leitura para pontuar as dicas.

## Como selecionar
- Priorize dicas com `impacto` alto e `evidencia` alta.
- Trate `novidade` como diferencial, nao como substituto de prova.
- So mantenha ideias com `risco_cliche` alto quando houver evidencias fortes e um mecanismo claro.
- Prefira termos que o professor realmente usa: `indisciplina`, `gestao de sala`, `rotina`, `autoridade`, `protocolos`, `scripts`, `primeiros 90 dias`.

## Sobre o relatorio do Gemini
- A base foi preparada para receber os 60 itens do relatorio compartilhado.
- O link compartilhado nao ficou acessivel de forma publica neste ambiente.
- Assim que o texto bruto for colado ou exportado, os slots podem ser preenchidos e classificados.

## Fontes iniciais ja registradas
- YouTube comment signals.
- Google Trends exports.
- Sales pages and copy texts.
- Market summary do projeto.
- Relatorio compartilhado do Gemini, pendente de importacao.
"""
    path.write_text(content, encoding="utf-8")


def sync_root(root: Path) -> None:
    rows = make_rows()
    write_csv(rel_path(root, CSV_REL), rows)
    create_workbook(rel_path(root, XLSX_REL), rows)
    write_readme(rel_path(root, README_REL))


def main() -> None:
    for root in TARGET_ROOTS:
        sync_root(root)

    print("Created:")
    for root in TARGET_ROOTS:
        print(rel_path(root, CSV_REL))
        print(rel_path(root, XLSX_REL))
        print(rel_path(root, README_REL))


if __name__ == "__main__":
    main()
