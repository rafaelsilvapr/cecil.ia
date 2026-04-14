from __future__ import annotations

import csv
from copy import copy
from pathlib import Path
from statistics import mean

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.workbook.properties import CalcProperties


BASE_DIR = Path("/Users/rafaelrodriguesdasilva/Documents/Agentes - Antigravity/Prof. Rafael")
DRIVE_BASE_DIR = Path(
    "/Users/rafaelrodriguesdasilva/Library/CloudStorage/GoogleDrive-rafaelsilva.pr@gmail.com/Meu Drive/Empreendedor/Professor Rafael/Prof. Rafael"
)

TARGET_ROOTS = [BASE_DIR, DRIVE_BASE_DIR]

CSV_REL = Path("output/spreadsheet/mapa_titulos_ebooks_gestao_sala_2026-04-08.csv")
XLSX_REL = Path("output/spreadsheet/mapa_titulos_ebooks_gestao_sala_2026-04-08.xlsx")
README_REL = Path("Content/base_de_dados/ebooks_gestao_sala/planejamento/README.md")

TITLE_ROWS = [
    {
        "id": "T01",
        "cluster": "indisciplina",
        "titulo_provisorio": "Indisciplina na Sala de Aula: Como Recuperar o Controle sem Gritar",
        "subtitulo_ou_promessa": "Estratégias diretas para retomar autoridade, reduzir desgaste e conduzir a turma com mais firmeza.",
        "dor_principal": "Perda de controle, exaustao emocional e conflito frequente em sala.",
        "solucao_clara": "Recuperar autoridade com procedimentos, limites e linguagem pratica.",
        "publico_alvo": "professores da educacao basica",
        "evidencias": "Trends: indisciplina e controle de sala; YouTube: scripts, protocolos e limites; oferta: Hotmart, Clube de Autores e cursos com promessa de controle.",
        "palavras_chave": "indisciplina, controle, autoridade, sala de aula, limites",
        "score_demanda": 5,
        "score_clareza": 5,
        "score_diferenciacao": 4,
        "score_execucao": 5,
        "status": "prioritario",
        "proximo_passo": "Escrever outline e validar subtitulo.",
        "notas": "Melhor candidato para capa de dor forte e promessa imediata.",
    },
    {
        "id": "T02",
        "cluster": "rotina",
        "titulo_provisorio": "Gestão de sala de aula sem caos: 40 protocolos práticos para reduzir a indisciplina no dia a dia",
        "subtitulo_ou_promessa": "Rotina, regras e intervenções para professores que precisam de método prático e objetivo.",
        "dor_principal": "Falta de previsibilidade, bagunça cotidiana e desgaste com improviso.",
        "solucao_clara": "Criar rotina, contrato e protocolos simples para conduzir a turma com firmeza.",
        "publico_alvo": "professores que querem mais organização",
        "evidencias": "Trends: gestao de sala, rotina e disciplina escolar; YouTube: planejamento, quick wins e automacao; oferta: Nova Escola e Hotmart.",
        "palavras_chave": "gestão de sala, indisciplina, protocolos, rotina, controle",
        "score_demanda": 5,
        "score_clareza": 5,
        "score_diferenciacao": 4,
        "score_execucao": 4,
        "status": "prioritario",
        "proximo_passo": "Fechar promessa principal e mapa de capitulos.",
        "notas": "Título final escolhido para o ebook; muito forte para professor sobrecarregado que busca ordem prática.",
    },
    {
        "id": "T03",
        "cluster": "tempo_planejamento",
        "titulo_provisorio": "Planejamento de Aula em 30 Minutos: O Guia Pratico para Quem Nao Tem Tempo",
        "subtitulo_ou_promessa": "Como planejar com rapidez sem perder clareza, intencao pedagógica e segurança para entrar em sala.",
        "dor_principal": "Falta de tempo para planejar e sensação de estar sempre atrasado.",
        "solucao_clara": "Usar um modelo enxuto de planejamento com passos repetiveis.",
        "publico_alvo": "professores sem tempo para preparar aulas",
        "evidencias": "YouTube: dor de tempo, quick wins e eficiencia; Trends: planejamento e organizacao; oferta: materiais com promessa de praticidade.",
        "palavras_chave": "planejamento, 30 minutos, aula, rapidez, praticidade",
        "score_demanda": 4,
        "score_clareza": 5,
        "score_diferenciacao": 4,
        "score_execucao": 5,
        "status": "prioritario",
        "proximo_passo": "Testar subtitulo e sumario de 5 blocos.",
        "notas": "Excelente para um ebook curto, vendavel e muito util.",
    },
    {
        "id": "T04",
        "cluster": "inicio_carreira",
        "titulo_provisorio": "Professor Iniciante: Como Sobreviver aos Primeiros 90 Dias na Escola",
        "subtitulo_ou_promessa": "Um mapa pratico para começar com menos culpa, mais organização e decisões mais seguras.",
        "dor_principal": "Inseguranca, choque com a realidade e sobrecarga no comeco da carreira.",
        "solucao_clara": "Organizar os primeiros passos e reduzir erros comuns de iniciante.",
        "publico_alvo": "professores iniciantes e licenciandos",
        "evidencias": "Trends: inicio de carreira e sobrecarga; YouTube: primeiros 90 dias e sobrevivencia; oferta: cursos e guias para professor iniciante.",
        "palavras_chave": "professor iniciante, primeiros 90 dias, sobrevivencia, escola",
        "score_demanda": 4,
        "score_clareza": 4,
        "score_diferenciacao": 4,
        "score_execucao": 4,
        "status": "prioritario",
        "proximo_passo": "Estruturar por semanas ou fases.",
        "notas": "Recorte de mercado muito promissor se a copy ficar concreta.",
    },
    {
        "id": "T05",
        "cluster": "scripts",
        "titulo_provisorio": "Scripts Prontos para Conversas Dificeis com Alunos, Familias e Coordenação",
        "subtitulo_ou_promessa": "Modelos de fala, mensagens e protocolos para nao travar na hora do conflito.",
        "dor_principal": "Dificuldade de conversar em situacoes tensas sem escalar o problema.",
        "solucao_clara": "Oferecer scripts e respostas prontas para dialogos sensiveis.",
        "publico_alvo": "professores e coordenadores",
        "evidencias": "YouTube: scripts e conversas dificeis; Trends: comunicacao e conflito; oferta: materiais prontos e exemplos de fala.",
        "palavras_chave": "scripts, conversas dificeis, familia, coordenação, dialogo",
        "score_demanda": 4,
        "score_clareza": 5,
        "score_diferenciacao": 4,
        "score_execucao": 4,
        "status": "prioritario",
        "proximo_passo": "Escrever blocos por situacao real.",
        "notas": "Bom apelo comercial porque entrega uma utilidade tangivel.",
    },
    {
        "id": "T06",
        "cluster": "protocolos",
        "titulo_provisorio": "Protocolo da Sala de Aula: Celular, Limites e Conflitos na Pratica",
        "subtitulo_ou_promessa": "Como criar regras claras e responder a conflitos sem depender da improvisacao do dia.",
        "dor_principal": "Falta de protocolo claro para celular, limites e incidentes recorrentes.",
        "solucao_clara": "Criar um conjunto de procedimentos que guiem a resposta do professor.",
        "publico_alvo": "professores que querem respostas objetivas",
        "evidencias": "YouTube: protocol, celular e limites; Trends: disciplina escolar; oferta: cursos com regras e procedimentos.",
        "palavras_chave": "protocolo, celular, limites, conflitos, sala de aula",
        "score_demanda": 4,
        "score_clareza": 4,
        "score_diferenciacao": 4,
        "score_execucao": 4,
        "status": "prioritario",
        "proximo_passo": "Definir situacoes e respostas-padrao.",
        "notas": "Forte quando a escola vive conflitos com tecnologia e convivencia.",
    },
    {
        "id": "T07",
        "cluster": "disciplina",
        "titulo_provisorio": "Disciplina Escolar na Pratica: Regras Claras e Intervencoes Simples",
        "subtitulo_ou_promessa": "Um guia direto para lidar com comportamento e manter a turma andando sem drama.",
        "dor_principal": "Comportamento dificil e dificuldade de sustentar limites consistentes.",
        "solucao_clara": "Regras simples, intervenções objetivas e rotina de resposta.",
        "publico_alvo": "professores e orientadores",
        "evidencias": "Trends: disciplina escolar e indisciplina; oferta: Educamundo, ProfessorIdeal e Hotmart; YouTube: fala direta sobre comportamento.",
        "palavras_chave": "disciplina escolar, regras, intervencoes, comportamento",
        "score_demanda": 4,
        "score_clareza": 4,
        "score_diferenciacao": 3,
        "score_execucao": 4,
        "status": "backlog",
        "proximo_passo": "Comparar com os outros titulos de disciplina.",
        "notas": "Boa base de nicho, mas menos afiado do que indisciplina/controle.",
    },
    {
        "id": "T08",
        "cluster": "rotina",
        "titulo_provisorio": "Rotina que Funciona na Sala de Aula: Como Organizar a Turma com Leveza",
        "subtitulo_ou_promessa": "Passos práticos para dar previsibilidade ao dia sem transformar a sala em um ambiente rigido.",
        "dor_principal": "Caos operacional e desgaste por falta de sequencia clara.",
        "solucao_clara": "Estruturar rotina com previsibilidade, fluidez e combinados simples.",
        "publico_alvo": "professores que buscam mais leveza",
        "evidencias": "Trends: rotina e gestao de sala; YouTube: tempo, previsibilidade e quick wins; oferta: materiais práticos com rotina.",
        "palavras_chave": "rotina, organizacao, turma, previsibilidade, leveza",
        "score_demanda": 4,
        "score_clareza": 4,
        "score_diferenciacao": 3,
        "score_execucao": 4,
        "status": "backlog",
        "proximo_passo": "Reforcar a promessa com um mecanismo unico.",
        "notas": "Pode funcionar melhor como subtitulo de um livro maior.",
    },
    {
        "id": "T09",
        "cluster": "neurodivergencia",
        "titulo_provisorio": "TDAH em Sala de Aula: Como Adaptar Sem Perder a Rotina da Turma",
        "subtitulo_ou_promessa": "Adaptações praticas para incluir sem desmontar o andamento da sala.",
        "dor_principal": "Dificuldade de atender necessidades individuais sem perder o controle coletivo.",
        "solucao_clara": "Adaptar a condução da turma com apoio de rotina e combinados inclusivos.",
        "publico_alvo": "professores que atendem estudantes com TDAH",
        "evidencias": "YouTube: neurodivergencia e TDAH; demanda latente em comentarios; oferta: materiais de adaptacao e inclusão.",
        "palavras_chave": "TDAH, inclusao, adaptação, rotina, turma",
        "score_demanda": 3,
        "score_clareza": 4,
        "score_diferenciacao": 3,
        "score_execucao": 3,
        "status": "backlog",
        "proximo_passo": "Validar se entra melhor como capitulo ou livro isolado.",
        "notas": "Bom tema, mas exige cuidado para nao prometer demais.",
    },
    {
        "id": "T10",
        "cluster": "familia_escola",
        "titulo_provisorio": "Pais Dificeis e a Escola Real: Como Lidar com Interferencias sem Escalar o Conflito",
        "subtitulo_ou_promessa": "Ferramentas para conversar com familias, manter limites e proteger a energia da escola.",
        "dor_principal": "Interferencias de familia, ruído na comunicação e conflito com a escola.",
        "solucao_clara": "Estruturar respostas, limites e canais de conversa.",
        "publico_alvo": "professores, coordenação e gestão",
        "evidencias": "YouTube: familia-escola e conversas dificeis; Trends: comunicacao e convivencia; oferta: guias de mediação e relacionamento.",
        "palavras_chave": "pais dificies, familia, escola, interferencia, conflito",
        "score_demanda": 4,
        "score_clareza": 4,
        "score_diferenciacao": 3,
        "score_execucao": 4,
        "status": "backlog",
        "proximo_passo": "Comparar com scripts e protocolos antes de priorizar.",
        "notas": "Tema muito relevante, mas pode entrar como eixo secundario do livro principal.",
    },
]


HEADERS = [
    "id",
    "cluster",
    "titulo_provisorio",
    "subtitulo_ou_promessa",
    "dor_principal",
    "solucao_clara",
    "publico_alvo",
    "evidencias",
    "palavras_chave",
    "score_demanda",
    "score_clareza",
    "score_diferenciacao",
    "score_execucao",
    "score_total",
    "status",
    "proximo_passo",
    "notas",
]


def score_total(item: dict) -> float:
    return round(
        mean(
            [
                item["score_demanda"],
                item["score_clareza"],
                item["score_diferenciacao"],
                item["score_execucao"],
            ]
        ),
        1,
    )


def ensure_parent(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def target_path(root: Path, rel: Path) -> Path:
    return root / rel


def write_csv(rows: list[dict], path: Path) -> None:
    ensure_parent(path)
    with path.open("w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=HEADERS)
        writer.writeheader()
        writer.writerows(rows)


def apply_sheet_header_style(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin", color="D9E1F2"),
        right=Side(style="thin", color="D9E1F2"),
        top=Side(style="thin", color="D9E1F2"),
        bottom=Side(style="thin", color="D9E1F2"),
    )


def style_body_cell(cell) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)


def build_workbook(rows: list[dict], path: Path) -> None:
    ensure_parent(path)
    wb = Workbook()
    wb.calculation = CalcProperties(calcMode="auto", fullCalcOnLoad=True, forceFullCalc=True)

    ws = wb.active
    ws.title = "Titulos"
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A2"

    for row in rows:
        row["score_total"] = score_total(row)

    ws.append(HEADERS)
    for item in rows:
        ws.append([item.get(header, "") for header in HEADERS])

    for cell in ws[1]:
        apply_sheet_header_style(cell)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            style_body_cell(cell)

    # Formula in the workbook so the score updates when the user edits the scores.
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=14).value = f"=ROUND(AVERAGE(J{row_idx}:M{row_idx}),1)"

    widths = {
        "A": 7,
        "B": 18,
        "C": 42,
        "D": 42,
        "E": 32,
        "F": 34,
        "G": 24,
        "H": 48,
        "I": 34,
        "J": 12,
        "K": 12,
        "L": 12,
        "M": 12,
        "N": 12,
        "O": 14,
        "P": 28,
        "Q": 34,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 46

    table_ref = f"A1:Q{ws.max_row}"
    table = Table(displayName="TitulosTable", ref=table_ref)
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    ws.add_table(table)

    dv = DataValidation(
        type="list",
        formula1='"backlog,prioritario,outline,rascunho,copy,capa,publicado"',
        allow_blank=False,
    )
    ws.add_data_validation(dv)
    dv.add(f"O2:O{ws.max_row}")

    # System sheet.
    sys = wb.create_sheet("Sistema")
    sys.sheet_view.showGridLines = False
    sys.freeze_panes = "A4"
    sys.column_dimensions["A"].width = 20
    sys.column_dimensions["B"].width = 28
    sys.column_dimensions["C"].width = 28
    sys.column_dimensions["D"].width = 36
    sys.column_dimensions["E"].width = 26
    sys.column_dimensions["F"].width = 18
    sys.column_dimensions["G"].width = 36
    sys.column_dimensions["H"].width = 36
    sys.column_dimensions["I"].width = 36

    section_fill = PatternFill("solid", fgColor="D9EAF7")
    section_font = Font(bold=True, color="1F1F1F")

    sys["A1"] = "Sistema de Escrita"
    sys["A1"].font = Font(bold=True, size=14, color="1F1F1F")
    sys["A1"].fill = section_fill
    sys["A1"].alignment = Alignment(horizontal="left", vertical="center")
    sys.merge_cells("A1:E1")

    sys["A3"] = "Fluxo"
    sys["A3"].font = section_font
    sys["A3"].fill = section_fill
    sys.merge_cells("A3:E3")

    flux_headers = ["Etapa", "Objetivo", "Entrada", "Saida", "Criterio"]
    flux_rows = [
        ["Captura", "Registrar sinais do mercado", "Trends, YouTube, ofertas, comentarios", "Banco bruto de sinais", "Nada importante fica fora"],
        ["Agrupar", "Organizar as dores por tema", "Banco bruto", "Clusters como indisciplina, rotina e planejamento", "Temas repetidos ficam visiveis"],
        ["Nomear", "Traduzir dor em promessa", "Clusters", "Titulos provisorios", "Titulo aponta um problema real"],
        ["Pontuar", "Escolher o que vale escrever agora", "Titulos provisorios", "Score e prioridade", "Score alto entra no funil"],
        ["Outline", "Transformar titulo em estrutura", "Titulo escolhido", "Mapa de capitulos", "Existe caminho claro de escrita"],
        ["Copy e capa", "Fechar promessa e embalagem", "Outline e evidencias", "Copy da pagina e da capa", "A capa confirma a dor e a promessa"],
    ]
    for col, value in enumerate(flux_headers, start=1):
        sys.cell(row=4, column=col, value=value)
        apply_sheet_header_style(sys.cell(row=4, column=col))
    for row_idx, row in enumerate(flux_rows, start=5):
        for col_idx, value in enumerate(row, start=1):
            cell = sys.cell(row=row_idx, column=col_idx, value=value)
            style_body_cell(cell)

    flux_table = Table(displayName="FluxoTable", ref="A4:E10")
    flux_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sys.add_table(flux_table)

    sys["A12"] = "Modelo de Registro"
    sys["A12"].font = section_font
    sys["A12"].fill = section_fill
    sys.merge_cells("A12:C12")

    brief_headers = ["Campo", "Como preencher", "Exemplo"]
    brief_rows = [
        ["cluster", "Tema macro da dor", "indisciplina"],
        ["titulo_provisorio", "Nome que vende a solucao", "Gestao de Sala de Aula sem Caos"],
        ["subtitulo_ou_promessa", "Resultado prometido", "Rotina, regras e estrategias que funcionam"],
        ["dor_principal", "Problema central do leitor", "Falta de previsibilidade e desgaste"],
        ["solucao_clara", "Como o livro resolve", "Procedimentos simples e repetiveis"],
        ["evidencias", "Sinais de mercado", "Trends, YouTube, ofertas e comentarios"],
        ["score_total", "Media dos 4 scores", "4.5"],
        ["status", "Fase atual", "prioritario"],
    ]
    for col, value in enumerate(brief_headers, start=1):
        sys.cell(row=13, column=col, value=value)
        apply_sheet_header_style(sys.cell(row=13, column=col))
    for row_idx, row in enumerate(brief_rows, start=14):
        for col_idx, value in enumerate(row, start=1):
            cell = sys.cell(row=row_idx, column=col_idx, value=value)
            style_body_cell(cell)

    brief_table = Table(displayName="BriefTable", ref="A13:C21")
    brief_table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium4",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sys.add_table(brief_table)

    sys["E12"] = "Regras"
    sys["E12"].font = section_font
    sys["E12"].fill = section_fill
    sys.merge_cells("E12:I12")

    rules = [
        "Comece pela dor, nao pelo tema abstrato.",
        "Se o titulo nao aponta uma solucao clara, ele volta para backlog.",
        "A copy deve usar a linguagem que ja aparece em Trends, YouTube e ofertas.",
        "O livro precisa caber em uma promessa praticavel e verificavel.",
        "Use o score para priorizar, mas nao ignore o contexto e a intencao estrategica.",
    ]
    for idx, rule in enumerate(rules, start=13):
        cell = sys.cell(row=idx, column=5, value=f"- {rule}")
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.font = Font(color="404040")

    for row in range(13, 18):
        sys.row_dimensions[row].height = 32

    sys.freeze_panes = "A4"
    wb.remove(wb["Sheet"]) if "Sheet" in wb.sheetnames else None
    wb.active = 0

    wb.save(path)


def write_readme(path: Path) -> None:
    ensure_parent(path)
    text = """# Planejamento de Titulos e Escrita - Ebooks de Gestao de Sala de Aula

Este diretorio guarda a camada de decisao de titulo e o sistema de escrita para os proximos ebooks do projeto.

## Arquivos
- `output/spreadsheet/mapa_titulos_ebooks_gestao_sala_2026-04-08.csv`
- `output/spreadsheet/mapa_titulos_ebooks_gestao_sala_2026-04-08.xlsx`

## Como usar
1. Abra a aba `Titulos` e ordene mentalmente pelo `score_total`.
2. Olhe primeiro para os titulos com `status = prioritario`.
3. Registre sempre `dor_principal`, `solucao_clara` e `evidencias` antes de escrever o outline.
4. Se a ideia nao encaixar em uma dor real, ela volta para `backlog`.
5. Quando escolher um titulo, avance para `outline`, depois `copy` e depois `capa`.

## O que os scores significam
- `score_demanda`: intensidade da procura ou dor observada nos dados.
- `score_clareza`: quao explicita e buscavel e a promessa.
- `score_diferenciacao`: quao diferente a ideia soa no mercado.
- `score_execucao`: quao rapido conseguimos transformar isso em livro.

## Regra pratica
O livro nao nasce do tema. Ele nasce da dor, da linguagem do mercado e da promessa clara.

## Candidatos mais fortes agora
- Indisciplina na Sala de Aula: Como Recuperar o Controle sem Gritar
- Gestão de sala de aula sem caos: 40 protocolos práticos para reduzir a indisciplina no dia a dia
- Planejamento de Aula em 30 Minutos: O Guia Pratico para Quem Nao Tem Tempo
- Professor Iniciante: Como Sobreviver aos Primeiros 90 Dias na Escola
- Scripts Prontos para Conversas Dificeis com Alunos, Familias e Coordenação
"""
    path.write_text(text, encoding="utf-8")


def mirror_to_targets(rel: Path, writer) -> None:
    for root in TARGET_ROOTS:
        writer(target_path(root, rel))


def main() -> int:
    rows = [copy(item) for item in TITLE_ROWS]
    rows.sort(key=lambda item: (-score_total(item), item["titulo_provisorio"]))
    for item in rows:
        item["score_total"] = score_total(item)

    def write_csv_target(path: Path) -> None:
        write_csv(rows, path)

    def write_xlsx_target(path: Path) -> None:
        build_workbook(rows, path)

    def write_readme_target(path: Path) -> None:
        write_readme(path)

    mirror_to_targets(CSV_REL, write_csv_target)
    mirror_to_targets(XLSX_REL, write_xlsx_target)
    mirror_to_targets(README_REL, write_readme_target)

    print(f"wrote {len(rows)} title rows")
    for root in TARGET_ROOTS:
        print(root / CSV_REL)
        print(root / XLSX_REL)
        print(root / README_REL)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
